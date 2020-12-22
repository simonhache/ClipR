#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Dec 21 13:38:17 2020

@author: simon
"""

# TODO: Set up the path and check the files (number and name)
import os
import openpyxl as xl

def clearConsole():
    print("\n"*100)

baseDirectory = "/Users/simon/Desktop/Correction Simon /"
os.chdir(baseDirectory)

correctionName = "_corrige_express"

for f in os.listdir():    
    os.chdir(baseDirectory + f)
    
    studentName = f.split("_")[0]
    xlFileName = studentName + correctionName + ".xlsx"


# TODO: Open the excel file and console.txt
    allFiles = os.listdir()
    for file in allFiles:
        if file[-3:] == "txt":
            fileName = file
        
    # Load the excel file
    xlWorkBook = xl.load_workbook(xlFileName)
    xlSheet = xlWorkBook .worksheets[0]
    
    # Setup the correction
    questionCol = 1
    correctionCol = 2
    valueCol = 3
    nRows = xlSheet.max_row
    answer = [] 
    
    print(f"Correcting {studentName} ({fileName})")
    with open(fileName, "r") as f:
        # Loop into the excel file 1 row at a time
        for row in range(1, nRows+1):
            questionCell = xlSheet.cell(row = row, column = questionCol)
            nextQuestionCell = xlSheet.cell(row=row+1,column = questionCol)
            valueCell = xlSheet.cell(row = row, column = valueCol)
            correctionCell = xlSheet.cell(row = row, column = correctionCol)
            
            # If the row is one with a correction to be made
            if not valueCell.value:
                pass
    
            elif not isinstance(valueCell.value, int):
               print(f"{questionCell.value} : {correctionCell}/{valueCell.value}") 
               
            else:
                a = True
                slow = False
                cnt = 0
                
                # while the question is being answered 
                while a is True:
                    clearConsole()
                    print("\u0332".join(f"Question : {questionCell.value} /{valueCell.value}"))
                    print("\n" *5)
                    os.system('clear')
                    cnt += 1
                    answer.append("           " + f.readline() + "        ") 
                    
                    if cnt > 20: # If the file didnt found the question in the first 10 lines
                        slow = True
                        
                    if nextQuestionCell.value in answer[-1] or slow is True:
                        print("".join(answer))
                        
                        query = input("Is this the whole question? (y/n)  ")
                        
                        if query == "y":
                            a = False
                            
                        else:
                            slow = True
                clearConsole()
                correction = float("inf")
                
                while correction > valueCell.value:
                    print(f"{questionCell.value} /{valueCell.value} : \n {''.join(answer)} \n")
                    correction = input(f"How many points ? (/{valueCell.value}):  ")
                    try :
                        correction = float(correction)
                    except Exception:
                        print("Enter a number")
                    
                correctionCell.value = correction    
                
                xlWorkBook.save(xlFileName)
                
                positionCheck = input("Is the current position ok? (y/n): ") == "y"
                
                if positionCheck is True: 
                    resetAnswer = input("Delete previous answer? (y/n) : ") == "y"
                    if resetAnswer:
                        answer = [answer[-1]]  
                
                while positionCheck is False:
                    rewindPosition = input("Go back for how many lines ?  ")
                    if rewindPosition.isnumeric():
                        answer = answer[-int(rewindPosition):]
                        print("".join(answer))
                        positionCheck = input("Is it ok now? (y/n) :  ") == "y"
                
                
            
